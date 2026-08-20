#!/usr/bin/env python3
"""Validate the committed WFM OS workbook and its reproducible build contract."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import zipfile
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.utils.cell import range_boundaries
except ImportError as exc:  # pragma: no cover - dependency message is intentional
    raise SystemExit(
        "openpyxl 3.1.5 is required. Install it with: "
        "python3 -m pip install -r requirements-build.txt"
    ) from exc


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT / "01_Application" / "WFM_OS.xlsx"
DEFAULT_PROVENANCE = ROOT / "01_Application" / "BUILD_PROVENANCE.json"

REQUIRED_SHEETS = {
    "00_HOME",
    "01_CONTROL_CENTER",
    "02_DATA_QUALITY",
    "03_REFRESH_AUDIT",
    "10_STRATEGIC_PLAN",
    "20_FORECAST",
    "24_SCHEDULE_DESIGN",
    "30_INTRADAY",
    "40_PERFORMANCE",
    "41_EXECUTIVE",
    "99_BUILD_INFO",
}

CONFIG_TABLES = {
    "tblParameters": ROOT / "02_Configuration" / "parameters.csv",
    "tblSourceSystems": ROOT / "02_Configuration" / "source_systems.csv",
    "tblFieldMapping": ROOT / "02_Configuration" / "field_mapping.csv",
    "tblValueMapping": ROOT / "02_Configuration" / "value_mapping.csv",
    "tblActivities": ROOT / "02_Configuration" / "activities.csv",
    "tblQueueMapping": ROOT / "02_Configuration" / "queue_mapping.csv",
    "tblMetricRules": ROOT / "02_Configuration" / "metric_rules.csv",
}

FORBIDDEN_CELL_TERMS = (
    "allianz",
    "verint",
    "claude",
    "204021",
)


def fail(message: str) -> None:
    raise RuntimeError(message)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def csv_headers(path: Path) -> list[str]:
    with path.open(newline="", encoding="utf-8") as source:
        return next(csv.reader(source))


def workbook_tables(workbook) -> dict[str, tuple[object, object]]:
    found: dict[str, tuple[object, object]] = {}
    for worksheet in workbook.worksheets:
        for table in worksheet.tables.values():
            if table.displayName in found:
                fail(f"Duplicate table name: {table.displayName}")
            found[table.displayName] = (worksheet, table)
    return found


def table_headers(worksheet, table) -> list[str]:
    min_col, min_row, max_col, _ = range_boundaries(table.ref)
    return [
        str(worksheet.cell(min_row, column).value or "")
        for column in range(min_col, max_col + 1)
    ]


def build_info(worksheet) -> dict[str, object]:
    result: dict[str, object] = {}
    for row in range(14, worksheet.max_row + 1):
        field = worksheet.cell(row, 2).value
        if field:
            result[str(field)] = worksheet.cell(row, 3).value
    return result


def internal_link_target(target: str) -> str | None:
    if not target.startswith("#"):
        return None
    match = re.match(r"^#(?:'((?:[^']|'')+)'|([^!]+))!", target)
    if not match:
        fail(f"Malformed internal hyperlink: {target}")
    return (match.group(1) or match.group(2)).replace("''", "'")


def validate_package(path: Path) -> None:
    if path.suffix.lower() != ".xlsx":
        fail("The shell must be an .xlsx file; macro-enabled promotion has not passed.")
    with zipfile.ZipFile(path) as archive:
        corrupt = archive.testzip()
        if corrupt:
            fail(f"Corrupt OOXML member: {corrupt}")
        names = archive.namelist()
        if any("externalLinks" in name for name in names):
            fail("External workbook links are not allowed in the portable shell.")
        if any(name.endswith("vbaProject.bin") for name in names):
            fail("The .xlsx shell unexpectedly contains a VBA project.")


def validate_workbook(path: Path) -> dict[str, object]:
    validate_package(path)
    workbook = load_workbook(path, read_only=False, data_only=False, keep_links=True)
    try:
        missing_sheets = sorted(REQUIRED_SHEETS.difference(workbook.sheetnames))
        if missing_sheets:
            fail(f"Missing required sheets: {', '.join(missing_sheets)}")

        visible = [sheet for sheet in workbook.worksheets if sheet.sheet_state == "visible"]
        hidden = [sheet for sheet in workbook.worksheets if sheet.sheet_state != "visible"]
        if len(workbook.sheetnames) != 43 or len(visible) != 35 or len(hidden) != 8:
            fail(
                "Unexpected workbook surface: "
                f"{len(workbook.sheetnames)} sheets, {len(visible)} visible, {len(hidden)} hidden"
            )
        if workbook.active.title != "00_HOME":
            fail("00_HOME must be the active opening sheet.")
        if getattr(workbook, "_external_links", []):
            fail("External workbook links are not allowed in the portable shell.")

        tables = workbook_tables(workbook)
        for table_name, contract_path in CONFIG_TABLES.items():
            if table_name not in tables:
                fail(f"Missing configuration table: {table_name}")
            worksheet, table = tables[table_name]
            actual = table_headers(worksheet, table)
            expected = csv_headers(contract_path)
            if actual != expected:
                fail(
                    f"Header drift in {table_name}: expected {expected!r}, found {actual!r}"
                )

        formula_count = 0
        validation_count = 0
        hyperlink_count = 0
        forbidden_hits: list[str] = []
        for worksheet in workbook.worksheets:
            validation_count += len(worksheet.data_validations.dataValidation)
            for row in worksheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, str):
                        if value.startswith("="):
                            formula_count += 1
                        lowered = value.lower()
                        for term in FORBIDDEN_CELL_TERMS:
                            if term in lowered:
                                forbidden_hits.append(f"{worksheet.title}!{cell.coordinate}:{term}")
                    if cell.hyperlink:
                        hyperlink_count += 1
                        target = internal_link_target(cell.hyperlink.target)
                        if target and target not in workbook.sheetnames:
                            fail(
                                f"Broken internal hyperlink at {worksheet.title}!{cell.coordinate}: "
                                f"{cell.hyperlink.target}"
                            )
        if forbidden_hits:
            fail("Non-portable legacy term(s) found: " + ", ".join(forbidden_hits))
        if formula_count < 4:
            fail(f"Expected shell formulas are missing; found only {formula_count}.")
        if validation_count < 10:
            fail(f"Expected controlled-input validations are missing; found only {validation_count}.")
        if hyperlink_count < 40:
            fail(f"Expected workbook navigation is incomplete; found only {hyperlink_count} links.")

        info = build_info(workbook["99_BUILD_INFO"])
        required_info = {
            "Release",
            "Operational status",
            "Build date",
            "Git commit",
            "Power Query",
            "Power Pivot / DAX",
            "Python in Excel",
            "VBA",
            "Production data",
        }
        missing_info = sorted(required_info.difference(info))
        if missing_info:
            fail(f"BUILD_INFO is incomplete: {', '.join(missing_info)}")
        if info["Operational status"] != "NOT OPERATIONAL":
            fail("The current shell must remain clearly marked NOT OPERATIONAL.")
        for engine in ("Power Query", "Power Pivot / DAX", "Python in Excel", "VBA"):
            if info[engine] != "NOT EMBEDDED":
                fail(f"BUILD_INFO makes an unsupported engine claim for {engine}.")
        if info["Production data"] != "NONE":
            fail("The committed workbook must remain data-free.")

        return {
            "build_info": info,
            "sheets": len(workbook.sheetnames),
            "visible": len(visible),
            "hidden": len(hidden),
            "tables": len(tables),
            "formulas": formula_count,
            "validations": validation_count,
            "hyperlinks": hyperlink_count,
        }
    finally:
        workbook.close()


def validate_provenance(path: Path, info: dict[str, object], provenance_path: Path) -> None:
    with provenance_path.open(encoding="utf-8") as source:
        provenance = json.load(source)
    expected = {
        "artifact": path.name,
        "build_date": str(info["Build date"]),
        "git_commit": str(info["Git commit"]),
        "release": str(info["Release"]),
        "sha256": sha256(path),
    }
    if provenance != expected:
        fail(
            "Workbook provenance drift: "
            f"expected {expected!r}, found {provenance!r}"
        )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--provenance", type=Path, default=DEFAULT_PROVENANCE)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    workbook_path = args.workbook.resolve()
    results = validate_workbook(workbook_path)
    validate_provenance(
        workbook_path,
        results["build_info"],
        args.provenance.resolve(),
    )
    print(
        "Validated WFM_OS.xlsx · "
        f"sha256={sha256(workbook_path)} · "
        f"sheets={results['sheets']} ({results['visible']} visible/{results['hidden']} hidden) · "
        f"tables={results['tables']} · formulas={results['formulas']} · "
        f"validations={results['validations']} · hyperlinks={results['hyperlinks']}"
    )


if __name__ == "__main__":
    main()
