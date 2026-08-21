#!/usr/bin/env python3
"""Build the deterministic, data-free WFM OS Excel application shell.

This generator intentionally creates an .xlsx workbook. It does not claim to
embed Power Query, Power Pivot, DAX, Python in Excel, or VBA. Those engines must
be installed and validated in desktop Excel before an .xlsm release is made.
"""

from __future__ import annotations

import argparse
import os
import re
import subprocess
import tempfile
import zipfile
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable, Sequence

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.formatting.rule import CellIsRule, FormulaRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.table import Table
    from openpyxl.workbook.defined_name import DefinedName
except ImportError as exc:  # pragma: no cover - dependency message is intentional
    raise SystemExit(
        "openpyxl 3.1+ is required. Install it with: "
        "python3 -m pip install openpyxl==3.1.5"
    ) from exc


TOKENS = {
    "ink-900": "111827",
    "ink-700": "334155",
    "ink-500": "64748B",
    "shell-950": "081824",
    "shell-900": "0B2233",
    "pearl-050": "F6F7F9",
    "paper-000": "FFFFFF",
    "line-200": "D8DEE7",
    "line-100": "E9EDF2",
    "primary-700": "0B6670",
    "primary-500": "168A96",
    "primary-100": "DCEFF1",
    "copper-600": "A4663F",
    "copper-100": "F2E6DD",
    "success-600": "18794E",
    "success-100": "DDF3E8",
    "warning-600": "B7791F",
    "warning-100": "FFF1D6",
    "danger-600": "B83A4B",
    "danger-100": "FBE4E8",
    "info-600": "315E9B",
    "info-100": "E2EBF7",
}


@dataclass(frozen=True)
class Page:
    name: str
    title: str
    decision: str
    accent: str
    kpis: tuple[str, str, str, str]
    analysis_title: str
    action_title: str
    detail_headers: tuple[str, ...]


BUSINESS_PAGES: tuple[Page, ...] = (
    Page(
        "10_STRATEGIC_PLAN",
        "Strategic plan",
        "Do demand, workforce supply, and investment remain aligned across the planning horizon?",
        "copper-600",
        ("PLANNING HORIZON", "DEMAND GROWTH", "FUNDED FTE", "CAPACITY GAP"),
        "Demand, supply, and cost outlook",
        "Planning assumptions requiring a decision",
        ("Period", "Activity", "Scenario", "Demand", "Required FTE", "Funded FTE", "Gap FTE", "Decision status"),
    ),
    Page(
        "11_DEMAND_PLAN",
        "Demand plan",
        "What workload should the operation plan for by activity, channel, and interval?",
        "copper-600",
        ("PLANNED VOLUME", "PLANNED AHT", "BACKLOG HOURS", "DEMAND CHANGE"),
        "Demand profile by activity and channel",
        "Demand assumptions and exceptional events",
        ("Period", "Activity", "Channel", "Baseline volume", "Event impact", "Planned volume", "AHT seconds", "Status"),
    ),
    Page(
        "20_FORECAST",
        "Forecast",
        "Which approved forecast version should drive downstream workforce decisions?",
        "copper-600",
        ("APPROVED VOLUME", "FORECAST AHT", "BIAS", "APPROVAL STATE"),
        "Actual, baseline, and approved forecast",
        "Forecast versions awaiting review",
        ("Version", "Activity", "Channel", "Start", "End", "Volume", "AHT seconds", "Approval status"),
    ),
    Page(
        "21_FORECAST_ACCURACY",
        "Forecast accuracy",
        "Where is forecast error systematic enough to change the next planning cycle?",
        "copper-600",
        ("WAPE", "BIAS", "INTERVAL ACCURACY", "OUTLIER PERIODS"),
        "Error trend and bias by planning grain",
        "Material misses for root-cause review",
        ("Period", "Activity", "Channel", "Actual", "Forecast", "Absolute error", "Bias", "Root-cause status"),
    ),
    Page(
        "22_CAPACITY_PLAN",
        "Capacity plan",
        "How many productive and paid FTE are required to meet demand and service objectives?",
        "copper-600",
        ("REQUIRED FTE", "AVAILABLE FTE", "NET GAP", "SHRINKAGE"),
        "Requirement, supply, and gap by period",
        "Capacity risks and scenario decisions",
        ("Period", "Activity", "Scenario", "Required FTE", "Available FTE", "Gap FTE", "Risk", "Decision owner"),
    ),
    Page(
        "23_HIRING_PLAN",
        "Hiring plan",
        "Which hiring and training waves close the capacity gap at the required proficiency date?",
        "copper-600",
        ("HIRING NEED", "IN PIPELINE", "RAMPING FTE", "RESIDUAL GAP"),
        "Hiring funnel, training waves, and proficiency",
        "Workforce actions requiring commitment",
        ("Wave", "Activity", "Hire date", "Training start", "Proficiency date", "Planned heads", "Expected yield", "Status"),
    ),
    Page(
        "24_SCHEDULE_DESIGN",
        "Schedule design",
        "Which shift mix covers interval demand while honoring contracts and policy?",
        "primary-700",
        ("COVERAGE", "PAID FTE", "PRODUCTIVE FTE", "CONSTRAINT BREACHES"),
        "Interval coverage and shift-pattern fit",
        "Schedule patterns requiring review",
        ("Pattern", "Activity", "Site", "Start", "End", "Paid hours", "Coverage contribution", "Status"),
    ),
    Page(
        "25_LEAVE_PLAN",
        "Leave plan",
        "How much leave can be approved without breaching service or coverage thresholds?",
        "primary-700",
        ("LEAVE ALLOWANCE", "REQUESTED", "APPROVED", "REMAINING"),
        "Leave allowance and capacity impact",
        "Requests outside available allowance",
        ("Date", "Activity", "Allowance hours", "Requested hours", "Approved hours", "Remaining hours", "Impact", "Status"),
    ),
    Page(
        "30_INTRADAY",
        "Intraday control",
        "What has changed today, what is the service impact, and which action should happen next?",
        "primary-500",
        ("SERVICE LEVEL", "VOLUME VS PLAN", "AHT VS PLAN", "NET STAFFING"),
        "Actual versus plan by interval",
        "Live exceptions and recommended actions",
        ("Interval", "Activity", "Channel", "Forecast", "Actual", "Required FTE", "Present FTE", "Action status"),
    ),
    Page(
        "31_ATTENDANCE",
        "Attendance",
        "Which attendance exceptions need evidence, ownership, and a controlled decision?",
        "primary-500",
        ("SCHEDULED", "PRESENT", "ABSENT", "OPEN CASES"),
        "Scheduled presence and attendance outcome",
        "Attendance cases requiring a decision",
        ("Case key", "Business date", "Agent key", "Exception", "Minutes", "Evidence status", "Owner", "Decision"),
    ),
    Page(
        "32_ADHERENCE",
        "Adherence",
        "Where did actual activity materially diverge from the approved schedule?",
        "primary-500",
        ("ADHERENCE", "CONFORMANCE", "EXCEPTION MINUTES", "AGENTS AFFECTED"),
        "Scheduled and actual activity alignment",
        "Material adherence exceptions",
        ("Business date", "Agent key", "Scheduled activity", "Actual activity", "Minutes", "Tolerance", "Owner", "Status"),
    ),
    Page(
        "33_ACTION_LOG",
        "Operational actions",
        "Are actions owned, time-bound, and producing the intended operational outcome?",
        "primary-500",
        ("OPEN ACTIONS", "OVERDUE", "DUE TODAY", "EFFECTIVE ACTIONS"),
        "Action volume, aging, and effectiveness",
        "Current action register",
        ("Action key", "Opened at", "Module", "Severity", "Action", "Owner", "Due at", "Status"),
    ),
    Page(
        "40_PERFORMANCE",
        "Performance",
        "Which performance gaps are material, explainable, and actionable at team and agent level?",
        "shell-900",
        ("SERVICE OUTCOME", "PRODUCTIVITY", "QUALITY", "ATTENDANCE"),
        "Balanced operational performance",
        "Exceptions requiring coaching or process action",
        ("Period", "Activity", "Team", "Metric", "Actual", "Target", "Variance", "Action status"),
    ),
    Page(
        "41_EXECUTIVE",
        "Executive view",
        "Are service, cost, people, and risk moving in the direction committed to leadership?",
        "shell-900",
        ("SERVICE", "COST / OUTPUT", "WORKFORCE RISK", "PLAN CONFIDENCE"),
        "Enterprise WFM scorecard",
        "Decisions and risks for leadership",
        ("Period", "Dimension", "Measure", "Actual", "Target", "Trend", "Risk", "Executive decision"),
    ),
    Page(
        "42_BONUS_CONTROL",
        "Incentive control",
        "Is the proposed incentive result governed, approved, and safe to publish?",
        "shell-900",
        ("ELIGIBLE PEOPLE", "PROPOSED PAYOUT", "CONTROL FAILURES", "APPROVAL STATE"),
        "Policy outcome and payout distribution",
        "Blocking controls and approval exceptions",
        ("Run key", "Period", "Policy version", "Population", "Currency", "Proposed payout", "Control status", "Approval status"),
    ),
)


CONFIG_SHEETS = (
    ("59_PARAMETERS", "Parameters", "Set the portable deployment boundary and live/closed refresh context.", "tblParameters",
     ("Parameter", "Value", "DataType", "Description")),
    ("60_SOURCE_SYSTEMS", "Source systems", "Register enterprise source roles and generic adapters.", "tblSourceSystems",
     ("Profile", "SystemKey", "SourceRole", "Adapter", "RelativePath", "FilePattern", "Format", "Delimiter", "Encoding", "Culture", "Enabled")),
    ("61_FIELD_MAPPING", "Field mapping", "Map external columns into canonical WFM contracts.", "tblFieldMapping",
     ("Profile", "Adapter", "Entity", "SourceField", "CanonicalField", "DataType", "Required", "Enabled")),
    ("62_VALUE_MAPPING", "Value mapping", "Translate external codes into governed canonical values.", "tblValueMapping",
     ("Profile", "SystemKey", "Domain", "SourceValue", "CanonicalValue", "ValidFrom", "ValidTo", "Enabled")),
    ("63_PEOPLE", "People", "Govern the effective-dated workforce dimension without exposing source identities.", "tblPeople",
     ("Profile", "AgentKey", "EmployeeBusinessID", "DisplayName", "ActivityKey", "TeamKey", "ManagerKey", "SiteKey", "ContractHours", "ValidFrom", "ValidTo", "EmploymentStatus", "Enabled")),
    ("64_IDENTITY_MAPPING", "Identity mapping", "Resolve each source-system identity to one dated canonical agent key.", "tblIdentityMapping",
     ("Profile", "SystemKey", "ExternalAgentID", "AgentKey", "ValidFrom", "ValidTo", "Enabled")),
    ("65_ACTIVITIES", "Activities", "Define the stable operating activities used across the WFM cycle.", "tblActivities",
     ("Profile", "ActivityKey", "ActivityName", "ChannelKey", "Timezone", "ValidFrom", "ValidTo", "Enabled")),
    ("66_QUEUE_MAPPING", "Queue mapping", "Map source queues to canonical queues, activities, and channels.", "tblQueueMapping",
     ("Profile", "SystemKey", "SourceQueueID", "QueueKey", "QueueName", "ActivityKey", "ChannelKey", "ValidFrom", "ValidTo", "Enabled")),
    ("67_STATE_MAPPING", "State mapping", "Map source states to canonical presence and adherence behavior.", "tblStateMapping",
     ("Profile", "SystemKey", "SourceStateID", "StateKey", "StateName", "AdherenceClass", "PresentFlag", "ProductiveFlag", "ValidFrom", "ValidTo", "Enabled")),
    ("68_TARGETS", "Targets", "Govern operational targets by scope and effective period.", "tblTargets",
     ("TargetKey", "MetricKey", "ActivityKey", "Channel", "TargetValue", "Unit", "ValidFrom", "ValidTo", "Approved", "Owner")),
    ("69_CALENDAR_EVENTS", "Calendar events", "Record dated events and quantified demand or capacity effects.", "tblCalendarEvents",
     ("Profile", "EventKey", "EventName", "EventType", "ActivityKey", "ChannelKey", "StartAt", "EndAt", "ImpactType", "ImpactValue", "Approved", "Owner", "Notes")),
    ("70_SHIFT_RULES", "Shift rules", "Define contract, shift, break, and coverage constraints.", "tblShiftRules",
     ("RuleKey", "RuleType", "ScopeKey", "Value", "Unit", "ValidFrom", "ValidTo", "Enabled", "Owner", "Notes")),
    ("71_METRIC_RULES", "Metric rules", "Store effective-dated policy parameters; DAX remains the metric engine.", "tblMetricRules",
     ("Profile", "MetricKey", "RuleName", "Value", "Unit", "ValidFrom", "ValidTo", "Approved")),
    ("72_FORECAST_POLICIES", "Forecast policies", "Govern forecast method, grain, history, horizon, and seasonal behavior by operating scope.", "tblForecastPolicies",
     ("Profile", "PolicyKey", "ActivityKey", "ChannelKey", "Method", "Frequency", "HistoryPeriods", "HorizonPeriods", "SeasonLength", "MinimumHistory", "ValidFrom", "ValidTo", "Approved")),
    ("73_CAPACITY_POLICIES", "Capacity policies", "Govern queueing or workload assumptions before requirements can be approved.", "tblCapacityPolicies",
     ("Profile", "PolicyKey", "ActivityKey", "ChannelKey", "Method", "IntervalMinutes", "TargetServiceLevel", "AnswerTimeSeconds", "MaxOccupancy", "ShrinkagePct", "Concurrency", "FTEPerHead", "ValidFrom", "ValidTo", "Approved")),
)


INPUT_SHEETS = (
    ("80_FORECAST_OVERRIDES", "Forecast overrides", "Capture governed changes before approval into a forecast version.", "tblForecastOverrides",
     ("OverrideKey", "Profile", "ForecastVersionKey", "ActivityKey", "ChannelKey", "ForecastDate", "ForecastVolume", "ForecastAHTSeconds", "Reason", "RequestedBy", "RequestedAt", "ApprovalStatus", "ApprovedBy", "ApprovedAt")),
    ("81_ATTEND_DECISIONS", "Attendance decisions", "Store keyed decisions separately from refreshable attendance cases.", "tblAttendanceDecisions",
     ("CaseKey", "Decision", "DecisionReason", "Owner", "DecisionAt", "EvidenceReference", "Notes")),
    ("82_ACTION_INPUT", "Action input", "Maintain the controlled operational action register.", "tblActionInput",
     ("ActionKey", "OpenedAt", "Module", "Severity", "Action", "Owner", "DueAt", "Status", "Outcome", "ClosedAt")),
    ("83_SCENARIO_INPUTS", "Scenario inputs", "Define versioned assumptions for planning and simulation.", "tblScenarioInputs",
     ("ScenarioKey", "Profile", "ScenarioName", "ForecastPolicyKey", "CapacityPolicyKey", "ActivityKey", "ChannelKey", "StartDate", "EndDate", "VolumeChangePct", "AHTChangePct", "ShrinkagePct", "Notes", "Status")),
    ("84_CLOSE_DAY", "Close day", "Approve one business date for a complete, reconciled, append-only operational snapshot.", "tblCloseDayInput",
     ("CloseKey", "Profile", "BusinessDate", "RequestedBy", "RequestedAt", "ApprovalStatus", "ApprovedBy", "ApprovedAt", "SnapshotStatus", "SnapshotAt", "SourceRunKey", "Notes")),
    ("85_FORECAST_APPROVAL", "Forecast approval", "Publish stable forecast versions only after analytical review and reconciliation.", "tblForecastVersions",
     ("ForecastRowKey", "Profile", "ForecastVersionKey", "ApprovalStatus", "Scenario", "Method", "ActivityKey", "ChannelKey", "IntervalStart", "ForecastVolume", "ForecastAHTSeconds", "CreatedAt", "ApprovedAt", "ApprovedBy", "SourceRunKey", "Notes")),
    ("86_REQUIREMENT_APPROVAL", "Requirement approval", "Publish capacity candidates into the operational staffing-requirement contract.", "tblRequirementApprovals",
     ("RequirementKey", "Profile", "ForecastVersionKey", "CapacityPolicyKey", "ApprovalStatus", "IntervalStart", "ActivityKey", "RequiredFTE", "PaidFTE", "RequiredHeads", "ShrinkagePct", "RequirementVersion", "ApprovedAt", "ApprovedBy", "SourceRunKey", "Notes")),
)


SOURCE_ROLES = (
    ("BLANK_DEPLOYMENT", "PEOPLE", "PeopleSource", "GenericDelimited", "03_Data/01_People", "*.*", "Auto", "comma", "65001", "en-US", False),
    ("BLANK_DEPLOYMENT", "CONTACTS", "ContactPlatform", "GenericDelimited", "03_Data/02_Contacts", "*.*", "Auto", "comma", "65001", "en-US", False),
    ("BLANK_DEPLOYMENT", "WORK_ITEMS", "WorkItemPlatform", "GenericDelimited", "03_Data/03_Work_Items", "*.*", "Auto", "comma", "65001", "en-US", False),
    ("BLANK_DEPLOYMENT", "AGENT_EVENTS", "AgentEventSource", "GenericDelimited", "03_Data/04_Agent_Events", "*.*", "Auto", "comma", "65001", "en-US", False),
    ("BLANK_DEPLOYMENT", "LOGIN_SESSIONS", "LoginSource", "GenericDelimited", "03_Data/05_Login_Sessions", "*.*", "Auto", "comma", "65001", "en-US", False),
    ("BLANK_DEPLOYMENT", "SCHEDULES", "SchedulingTool", "GenericDelimited", "03_Data/06_Schedules", "*.*", "Auto", "comma", "65001", "en-US", False),
    ("BLANK_DEPLOYMENT", "ABSENCE_LEAVE", "AbsenceLeaveSource", "GenericDelimited", "03_Data/07_Absence_Leave", "*.*", "Auto", "comma", "65001", "en-US", False),
    ("BLANK_DEPLOYMENT", "QUALITY", "QualitySource", "GenericDelimited", "03_Data/08_Quality", "*.*", "Auto", "comma", "65001", "en-US", False),
    ("BLANK_DEPLOYMENT", "FORECASTS", "ForecastSource", "GenericDelimited", "03_Data/09_Forecasts", "*.*", "Auto", "comma", "65001", "en-US", False),
    ("BLANK_DEPLOYMENT", "STAFFING_REQUIREMENTS", "StaffingRequirementSource", "GenericDelimited", "03_Data/10_Staffing_Requirements", "*.*", "Auto", "comma", "65001", "en-US", False),
)


ADAPTER_CONTRACTS = {
    "Contact": (
        "SourceContactID", "ContactStart", "ContactEnd", "ChannelExternalID",
        "QueueExternalID", "AgentExternalID", "Direction", "Outcome",
        "WaitSeconds", "TalkSeconds", "HoldSeconds", "AfterContactSeconds"
    ),
    "WorkItem": (
        "SourceWorkItemID", "CreatedAt", "CompletedAt", "ChannelExternalID", "WorkTypeExternalID", "ActivityExternalID", "AgentExternalID", "StatusExternalID", "HandlingSeconds", "SlaDeadline"
    ),
    "AgentEvent": ("SourceAgentEventID", "AgentExternalID", "EventStart", "EventEnd", "StateExternalID"),
    "LoginSession": ("SourceLoginSessionID", "AgentExternalID", "LoginAt", "LogoutAt"),
    "ScheduleSegment": ("SourceScheduleSegmentID", "AgentExternalID", "ScheduledStart", "ScheduledEnd", "ActivityExternalID", "ScheduleTypeExternalID", "PaidFlag", "ProductiveFlag"),
    "StaffingRequirement": ("SourceRequirementID", "IntervalStart", "ActivityExternalID", "RequiredFTE", "RequirementVersion", "ApprovedFlag"),
    "Forecast": ("ForecastVersionKey", "ApprovalStatus", "Scenario", "Method", "ActivityKey", "ChannelKey", "IntervalStart", "ForecastVolume", "ForecastAHTSeconds", "CreatedAt", "ApprovedAt", "ApprovedBy", "SourceRunKey"),
}


OPTIONAL_ADAPTER_FIELDS = {
    "ContactEnd",
    "AgentExternalID",
    "TalkSeconds",
    "HoldSeconds",
    "AfterContactSeconds",
    "CompletedAt",
    "AgentKey",
    "ApprovedAt",
    "RequirementVersion",
    "Scenario",
    "ApprovedBy",
    "SourceRunKey",
}


def adapter_data_type(field_name: str) -> str:
    if field_name.endswith(("At", "Start", "End")):
        return "datetime"
    if field_name.endswith(("Seconds", "Hours", "Volume", "AHT", "FTE")):
        return "number"
    if field_name.endswith("Flag"):
        return "boolean"
    return "text"


def fill(token: str) -> PatternFill:
    return PatternFill("solid", fgColor=TOKENS[token])


def side(token: str, style: str = "thin") -> Side:
    return Side(style=style, color=TOKENS[token])


def font(size: float = 10, color: str = "ink-700", bold: bool = False) -> Font:
    return Font(name="Aptos", size=size, color=TOKENS[color], bold=bold)


def clean_table_name(sheet_name: str) -> str:
    return "tbl" + "".join(part.title() for part in sheet_name.split("_") if not part.isdigit())


def set_canvas(ws, tab_color: str = "ink-500") -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 100
    ws.sheet_properties.tabColor = TOKENS[tab_color]
    ws.freeze_panes = "B7"
    ws.sheet_format.defaultRowHeight = 16
    ws.column_dimensions["A"].width = 2.5
    ws.column_dimensions["R"].width = 2.5
    for col in range(2, 18):
        ws.column_dimensions[get_column_letter(col)].width = 12
    ws.row_dimensions[1].height = 8
    for row in range(1, 120):
        for col in range(1, 19):
            ws.cell(row, col).fill = fill("pearl-050")


def merge_style(ws, cell_range: str, value: object, *, cell_fill: str = "paper-000", cell_font: Font | None = None,
                alignment: Alignment | None = None, border: Border | None = None) -> None:
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = value
    cell.fill = fill(cell_fill)
    cell.font = cell_font or font()
    cell.alignment = alignment or Alignment(vertical="center")
    if border:
        cell.border = border


def apply_fill_to_range(ws, min_row: int, max_row: int, min_col: int, max_col: int, token: str) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.fill = fill(token)


def set_internal_link(cell, label: str, target: str, *, dark: bool = True) -> None:
    cell.value = label
    cell.hyperlink = f"#'{target}'!B2"
    cell.font = Font(
        name="Aptos",
        size=8.5,
        bold=True,
        color=TOKENS["paper-000" if dark else "primary-700"],
        underline="single",
    )
    cell.alignment = Alignment(horizontal="center", vertical="center")


def page_shell(ws, page_name: str, title: str, decision: str, accent: str, previous_name: str, next_name: str) -> None:
    set_canvas(ws, accent)
    apply_fill_to_range(ws, 2, 3, 2, 17, "shell-950")
    merge_style(ws, "B2:D3", "WFM OS", cell_fill="shell-950", cell_font=font(18, "paper-000", True))
    merge_style(ws, "E2:J3", title.upper(), cell_fill="shell-950", cell_font=font(9, "paper-000", True))
    merge_style(ws, "K2:M3", "SHELL · NOT OPERATIONAL", cell_fill="shell-900", cell_font=font(8.5, "warning-100", True), alignment=Alignment(horizontal="center", vertical="center"))
    for col, label, target in ((14, "HOME", "00_HOME"), (15, "PREVIOUS", previous_name), (16, "NEXT", next_name), (17, "BUILD", "99_BUILD_INFO")):
        set_internal_link(ws.cell(2, col), label, target)
        ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
    for col in range(2, 18):
        ws.cell(3, col).border = Border(bottom=side("primary-500", "medium"))
    merge_style(ws, "B4:Q4", title, cell_fill="pearl-050", cell_font=font(18, "ink-900", True), alignment=Alignment(vertical="bottom"))
    merge_style(ws, "B5:Q5", decision, cell_fill="pearl-050", cell_font=font(9, "ink-500"), alignment=Alignment(vertical="top", wrap_text=True))
    ws.row_dimensions[4].height = 26
    ws.row_dimensions[5].height = 22
    ws.auto_filter.ref = None
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.outlinePr.summaryBelow = True


def control_rail(ws) -> None:
    blocks = (("B7:E9", "PROFILE\nBLANK_DEPLOYMENT"), ("F7:I9", "AS-OF DATE\nNot set"), ("J7:M9", "ACTIVITY\nAll configured"), ("N7:Q9", "SCENARIO\nApproved"))
    for cell_range, text in blocks:
        merge_style(
            ws,
            cell_range,
            text,
            cell_fill="paper-000",
            cell_font=font(9, "ink-700", True),
            alignment=Alignment(vertical="center", wrap_text=True, indent=1),
            border=Border(left=side("line-200"), right=side("line-200"), top=side("line-200"), bottom=side("line-200")),
        )


def kpi_cards(ws, labels: Sequence[str]) -> None:
    ranges = ((2, 5), (6, 9), (10, 13), (14, 17))
    for (start_col, end_col), label in zip(ranges, labels):
        for row in range(11, 16):
            for col in range(start_col, end_col + 1):
                ws.cell(row, col).fill = fill("paper-000")
        ws.merge_cells(start_row=11, start_column=start_col, end_row=11, end_column=end_col)
        ws.merge_cells(start_row=12, start_column=start_col, end_row=14, end_column=end_col)
        ws.merge_cells(start_row=15, start_column=start_col, end_row=15, end_column=end_col)
        label_cell = ws.cell(11, start_col)
        label_cell.value = label
        label_cell.font = font(8.5, "ink-500", True)
        label_cell.alignment = Alignment(vertical="center", indent=1)
        value_cell = ws.cell(12, start_col)
        value_cell.value = "—"
        value_cell.font = font(24, "ink-900", True)
        value_cell.alignment = Alignment(vertical="center", indent=1)
        status_cell = ws.cell(15, start_col)
        status_cell.value = "AWAITING VALIDATED DATA"
        status_cell.font = font(8, "ink-500", True)
        status_cell.alignment = Alignment(vertical="center", indent=1)
        for row in range(11, 16):
            ws.cell(row, start_col).border = Border(left=side("line-200", "medium"))


def section_box(ws, top: int, bottom: int, title: str, message: str, accent: str = "line-200") -> None:
    apply_fill_to_range(ws, top, bottom, 2, 17, "paper-000")
    ws.merge_cells(start_row=top, start_column=2, end_row=top, end_column=17)
    heading = ws.cell(top, 2)
    heading.value = title.upper()
    heading.font = font(10, "ink-700", True)
    heading.alignment = Alignment(vertical="center", indent=1)
    heading.border = Border(bottom=side("line-200", "medium"))
    ws.merge_cells(start_row=top + 2, start_column=3, end_row=bottom - 1, end_column=16)
    body = ws.cell(top + 2, 3)
    body.value = message
    body.font = font(10, "ink-500")
    body.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in range(top, bottom + 1):
        ws.cell(row, 2).border = Border(left=side(accent, "medium"))


def style_table(ws, start_row: int, start_col: int, headers: Sequence[str], rows: Sequence[Sequence[object]], table_name: str,
                *, input_table: bool = False, max_input_row: int = 500) -> Table:
    for index, header in enumerate(headers, start=start_col):
        cell = ws.cell(start_row, index, header)
        cell.font = font(9, "ink-700", True)
        cell.fill = fill("pearl-050")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = Border(bottom=side("ink-500", "medium"))
    normalized_rows = list(rows) or [[None] * len(headers)]
    for offset, row_data in enumerate(normalized_rows, start=1):
        for col_offset, value in enumerate(row_data):
            cell = ws.cell(start_row + offset, start_col + col_offset, value)
            cell.font = font(9, "ink-700")
            cell.fill = fill("primary-100" if input_table else "paper-000")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=side("line-100"))
            if input_table:
                cell.protection = Protection(locked=False)
        if input_table:
            ws.cell(start_row + offset, start_col).border = Border(left=side("primary-500", "medium"), bottom=side("line-100"))
    end_row = start_row + len(normalized_rows)
    end_col = start_col + len(headers) - 1
    table = Table(displayName=table_name, ref=f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}")
    ws.add_table(table)
    ws.auto_filter.ref = None
    ws.row_dimensions[start_row].height = 28
    if input_table:
        for row in range(end_row + 1, max_input_row + 1):
            for col in range(start_col, end_col + 1):
                ws.cell(row, col).fill = fill("pearl-050")
                ws.cell(row, col).protection = Protection(locked=False)
    return table


def business_page(ws, page: Page, previous_name: str, next_name: str) -> None:
    page_shell(ws, page.name, page.title, page.decision, page.accent, previous_name, next_name)
    control_rail(ws)
    kpi_cards(ws, page.kpis)
    section_box(
        ws,
        17,
        29,
        page.analysis_title,
        "No validated data is available. Complete setup, pass Data Quality, then refresh the approved release.",
        page.accent,
    )
    ws.merge_cells("B31:Q31")
    ws["B31"] = page.action_title.upper()
    ws["B31"].font = font(10, "ink-700", True)
    ws["B31"].fill = fill("pearl-050")
    style_table(ws, 33, 2, page.detail_headers, [], clean_table_name(page.name))
    ws.auto_filter.ref = None


def add_home(wb: Workbook, version: str) -> None:
    ws = wb.create_sheet("00_HOME")
    set_canvas(ws, "primary-700")
    ws.freeze_panes = "B7"
    apply_fill_to_range(ws, 2, 3, 2, 17, "shell-950")
    merge_style(ws, "B2:E3", "WFM OS", cell_fill="shell-950", cell_font=font(20, "paper-000", True))
    merge_style(ws, "F2:L3", "UNIVERSAL WORKFORCE MANAGEMENT", cell_fill="shell-950", cell_font=font(9, "paper-000", True))
    merge_style(ws, "M2:Q3", f"APPLICATION SHELL · {version}", cell_fill="shell-900", cell_font=font(8.5, "warning-100", True), alignment=Alignment(horizontal="center", vertical="center"))
    for col in range(2, 18):
        ws.cell(3, col).border = Border(bottom=side("primary-500", "medium"))
    merge_style(ws, "B4:Q4", "Command center", cell_fill="pearl-050", cell_font=font(18, "ink-900", True), alignment=Alignment(vertical="bottom"))
    merge_style(ws, "B5:Q5", "One governed view of demand, people, service, risk, and action across the complete WFM cycle.", cell_fill="pearl-050", cell_font=font(9, "ink-500"))
    cards = (
        ("B7:E10", "DEPLOYMENT", '=IF(COUNTIF(tblSourceSystems[Enabled],TRUE)>0,"CONFIGURED","NOT CONFIGURED")', "info-600"),
        ("F7:I10", "DATA QUALITY", '=IF(COUNTIF(tblDQChecks[Status],"BLOCKING")>0,"BLOCKED","NOT VALIDATED")', "info-600"),
        ("J7:M10", "LAST REFRESH", '=IF(COUNTA(tblRefreshAudit[CompletedAt])=0,"NEVER","SEE AUDIT")', "info-600"),
        ("N7:Q10", "RELEASE STATE", '="NOT OPERATIONAL"', "warning-600"),
    )
    for cell_range, label, formula, accent in cards:
        start, end = cell_range.split(":")
        start_col = ws[start].column
        end_col = ws[end].column
        start_row = ws[start].row
        end_row = ws[end].row
        apply_fill_to_range(ws, start_row, end_row, start_col, end_col, "paper-000")
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
        ws.merge_cells(start_row=start_row + 1, start_column=start_col, end_row=end_row, end_column=end_col)
        ws.cell(start_row, start_col, label).font = font(8.5, "ink-500", True)
        ws.cell(start_row, start_col).alignment = Alignment(indent=1, vertical="center")
        ws.cell(start_row + 1, start_col, formula).font = font(15, "ink-900", True)
        ws.cell(start_row + 1, start_col).alignment = Alignment(indent=1, vertical="center", wrap_text=True)
        for row in range(start_row, end_row + 1):
            ws.cell(row, start_col).border = Border(left=side(accent, "medium"))
    ws.merge_cells("B12:Q12")
    ws["B12"] = "WORKFORCE MANAGEMENT CYCLE"
    ws["B12"].font = font(10, "ink-700", True)
    ws["B12"].fill = fill("pearl-050")
    module_groups = (
        ("B14:F21", "PLAN", "Strategy · demand · forecast\ncapacity · hiring", "10_STRATEGIC_PLAN", "copper-600"),
        ("G14:L21", "OPERATE", "Schedule · leave · intraday\nattendance · adherence · actions", "24_SCHEDULE_DESIGN", "primary-700"),
        ("M14:Q21", "IMPROVE", "Performance · executive\nincentive governance", "40_PERFORMANCE", "shell-900"),
    )
    for cell_range, label, detail, target, accent in module_groups:
        start, end = cell_range.split(":")
        sc, ec, sr, er = ws[start].column, ws[end].column, ws[start].row, ws[end].row
        apply_fill_to_range(ws, sr, er, sc, ec, "paper-000")
        ws.merge_cells(start_row=sr, start_column=sc, end_row=sr + 1, end_column=ec)
        ws.merge_cells(start_row=sr + 2, start_column=sc, end_row=er - 1, end_column=ec)
        ws.merge_cells(start_row=er, start_column=sc, end_row=er, end_column=ec)
        title_cell = ws.cell(sr, sc, label)
        title_cell.font = font(14, "ink-900", True)
        title_cell.alignment = Alignment(indent=1, vertical="center")
        detail_cell = ws.cell(sr + 2, sc, detail)
        detail_cell.font = font(10, "ink-500")
        detail_cell.alignment = Alignment(indent=1, vertical="center", wrap_text=True)
        link_cell = ws.cell(er, sc)
        set_internal_link(link_cell, "OPEN MODULE  →", target, dark=False)
        link_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        for row in range(sr, er + 1):
            ws.cell(row, sc).border = Border(left=side(accent, "medium"))
    ws.merge_cells("B24:Q24")
    ws["B24"] = "START HERE"
    ws["B24"].font = font(10, "ink-700", True)
    steps = (
        ("B26:E31", "01", "Configure", "Register a deployment profile and source roles.", "01_CONTROL_CENTER"),
        ("F26:I31", "02", "Map", "Map external fields and values to canonical contracts.", "61_FIELD_MAPPING"),
        ("J26:M31", "03", "Validate", "Resolve every blocking quality and reconciliation check.", "02_DATA_QUALITY"),
        ("N26:Q31", "04", "Operate", "Use decisions only after the release gate is approved.", "99_BUILD_INFO"),
    )
    for cell_range, number, label, detail, target in steps:
        start, end = cell_range.split(":")
        sc, ec, sr, er = ws[start].column, ws[end].column, ws[start].row, ws[end].row
        apply_fill_to_range(ws, sr, er, sc, ec, "paper-000")
        ws.merge_cells(start_row=sr, start_column=sc, end_row=sr, end_column=ec)
        ws.merge_cells(start_row=sr + 1, start_column=sc, end_row=er - 1, end_column=ec)
        ws.merge_cells(start_row=er, start_column=sc, end_row=er, end_column=ec)
        ws.cell(sr, sc, f"{number}  {label.upper()}").font = font(9, "ink-700", True)
        ws.cell(sr, sc).alignment = Alignment(indent=1, vertical="center")
        ws.cell(sr + 1, sc, detail).font = font(9, "ink-500")
        ws.cell(sr + 1, sc).alignment = Alignment(indent=1, vertical="center", wrap_text=True)
        set_internal_link(ws.cell(er, sc), "OPEN  →", target, dark=False)
        ws.cell(er, sc).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    merge_style(ws, "B34:Q37", "This release is not approved for operational use. Review BUILD_INFO and complete the release gate first.", cell_fill="warning-100", cell_font=font(9, "warning-600", True), alignment=Alignment(vertical="center", wrap_text=True, indent=1), border=Border(left=side("warning-600", "medium")))


def add_control_center(wb: Workbook) -> None:
    ws = wb.create_sheet("01_CONTROL_CENTER")
    page_shell(ws, ws.title, "Control center", "Configure the deployment boundary and operating context before any source is enabled.", "info-600", "00_HOME", "02_DATA_QUALITY")
    section_box(ws, 7, 12, "Deployment readiness", "Configuration is intentionally disabled. Complete the controlled fields below, then validate source mappings and data quality.", "info-600")
    ws.merge_cells("B14:Q14")
    ws["B14"] = "CONTROLLED DEPLOYMENT SETTINGS"
    ws["B14"].font = font(10, "ink-700", True)
    settings = (
        ("Deployment profile", "BLANK_DEPLOYMENT", "Unique enterprise configuration profile."),
        ("As-of date", "", "Operational boundary between closed and live facts."),
        ("Business time zone", "UTC", "Canonical display and business-day context."),
        ("History months", 13, "Closed-fact history window."),
        ("Refresh mode", "VALIDATION", "Validation must pass before operational mode."),
        ("Source root", "..\\03_Data", "External runtime data; never commit production data."),
        ("Output root", "..\\04_Outputs", "Controlled publications and generated outputs."),
        ("Backup root", "..\\05_Backups", "Recoverable workbook snapshots."),
    )
    headers = ("Setting", "Value", "Purpose", "Control state")
    rows = tuple((label, value, purpose, "REQUIRED" if value in ("", "BLANK_DEPLOYMENT") else "DEFAULT") for label, value, purpose in settings)
    style_table(ws, 16, 2, headers, rows, "tblControlSettings", input_table=True)
    for row in range(17, 17 + len(rows)):
        ws.cell(row, 3).fill = fill("primary-100")
        ws.cell(row, 3).border = Border(left=side("primary-500", "medium"), bottom=side("line-100"))
    add_list_validation(ws, "C17", "C17", "PROFILE_LIST", "Choose a configured deployment profile.")
    add_list_validation(ws, "C19", "C19", "TIMEZONE_LIST", "Choose the canonical business time zone.")
    add_list_validation(ws, "C21", "C21", "REFRESH_MODE_LIST", "Use validation until all blocking checks pass.")
    ws["C18"].number_format = "yyyy-mm-dd"
    ws["C20"].number_format = "0"
    ws.auto_filter.ref = None


def add_data_quality(wb: Workbook) -> None:
    ws = wb.create_sheet("02_DATA_QUALITY")
    page_shell(ws, ws.title, "Data quality", "Which source or model issue blocks safe operational use, and who owns the resolution?", "info-600", "01_CONTROL_CENTER", "03_REFRESH_AUDIT")
    control_rail(ws)
    kpi_cards(ws, ("BLOCKING", "WARNINGS", "UNMAPPED", "LAST VALIDATION"))
    ws["B12"] = '=COUNTIF(tblDQChecks[Status],"BLOCKING")'
    ws["F12"] = '=COUNTIF(tblDQChecks[Status],"WARNING")'
    ws["J12"] = '=SUM(tblDQChecks[IssueCount])'
    ws["N12"] = "—"
    checks = (
        ("DQ-001", "Required source enabled", "Source", "At least one enabled source is required.", "NOT RUN", 0, "Deployment owner", "Configure source roles"),
        ("DQ-002", "Required fields mapped", "Schema", "Every required canonical field must have an enabled mapping.", "NOT RUN", 0, "Data owner", "Complete field mapping"),
        ("DQ-003", "Unknown values quarantined", "Mapping", "Unknown source values must remain visible and excluded from approval.", "NOT RUN", 0, "Operations owner", "Resolve value mapping"),
        ("DQ-004", "Canonical keys unique", "Model", "Dimension and fact keys must be unique at their declared grain.", "NOT RUN", 0, "Model owner", "Validate canonical build"),
        ("DQ-005", "Live and closed do not overlap", "Model", "No business date may exist in both refresh lanes.", "NOT RUN", 0, "Model owner", "Validate date boundary"),
        ("DQ-006", "Source totals reconcile", "Reconciliation", "Source totals must reconcile to canonical facts before approval.", "NOT RUN", 0, "WFM owner", "Run reconciliation"),
        ("DQ-007", "Forecast approvals are unique", "Planning", "Only one valid approved forecast row may exist at each interval, activity, and channel grain.", "NOT RUN", 0, "Planning owner", "Resolve dq_PlanningApprovals"),
        ("DQ-008", "Requirements are approved", "Planning", "Only evidence-backed, valid approved requirements may enter canonical staffing facts.", "NOT RUN", 0, "Capacity owner", "Resolve dq_PlanningApprovals"),
    )
    style_table(ws, 18, 2, ("CheckKey", "Check", "Domain", "Requirement", "Status", "IssueCount", "Owner", "NextAction"), checks, "tblDQChecks")
    add_list_validation(ws, "F19", "F500", "DQ_STATUS_LIST", "Choose the validation outcome.")
    ws.conditional_formatting.add("F19:F500", FormulaRule(formula=['F19="BLOCKING"'], fill=fill("danger-100"), font=font(9, "danger-600", True)))
    ws.conditional_formatting.add("F19:F500", FormulaRule(formula=['F19="WARNING"'], fill=fill("warning-100"), font=font(9, "warning-600", True)))
    ws.conditional_formatting.add("F19:F500", FormulaRule(formula=['F19="PASSED"'], fill=fill("success-100"), font=font(9, "success-600", True)))


def add_refresh_audit(wb: Workbook) -> None:
    ws = wb.create_sheet("03_REFRESH_AUDIT")
    page_shell(ws, ws.title, "Refresh audit", "Is the current view traceable to a complete, successful, and reconciled refresh?", "info-600", "02_DATA_QUALITY", "10_STRATEGIC_PLAN")
    section_box(ws, 7, 14, "Refresh history", "No refresh has been executed. The executable refresh controller will append immutable run records here in a later release.", "info-600")
    style_table(ws, 17, 2, ("RunKey", "Profile", "Mode", "StartedAt", "CompletedAt", "Status", "SourceFiles", "RowsRead", "RowsQuarantined", "ReconciliationStatus", "InitiatedBy"), [], "tblRefreshAudit")


def add_list_validation(ws, start: str, end: str, defined_name: str, prompt: str) -> None:
    dv = DataValidation(type="list", formula1=f"={defined_name}", allow_blank=True)
    dv.error = "Select a value from the governed list."
    dv.errorTitle = "Invalid controlled value"
    dv.prompt = prompt
    dv.promptTitle = "WFM OS control"
    dv.showErrorMessage = True
    dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add(f"{start}:{end}")


def config_or_input_sheet(wb: Workbook, spec, *, is_input: bool = False) -> None:
    name, title, decision, table_name, headers = spec
    ws = wb.create_sheet(name)
    accent = "primary-500" if is_input else "info-600"
    previous_name = wb.worksheets[-2].title
    next_name = "00_HOME"
    page_shell(ws, name, title, decision, accent, previous_name, next_name)
    if is_input:
        merge_style(ws, "B7:Q10", "CONTROLLED INPUT · Enter decisions only in blue cells. Keep stable keys and do not type beside refreshable query output.", cell_fill="primary-100", cell_font=font(9, "primary-700", True), alignment=Alignment(vertical="center", wrap_text=True, indent=1), border=Border(left=side("primary-500", "medium")))
    else:
        merge_style(ws, "B7:Q10", "CONFIGURATION · Changes affect the selected deployment profile after validation and approval. Unknown values must remain visible.", cell_fill="info-100", cell_font=font(9, "info-600", True), alignment=Alignment(vertical="center", wrap_text=True, indent=1), border=Border(left=side("info-600", "medium")))
    rows: Sequence[Sequence[object]] = []
    if name == "59_PARAMETERS":
        rows = (
            ("RootPath", "<SET ABSOLUTE ROOT PATH>", "text", "Absolute path of the cloned WFM OS folder"),
            ("EnterpriseProfile", "BLANK_DEPLOYMENT", "text", "Active deployment profile; replace before enabling sources"),
            ("AsOfDate", "", "date", "Operational date used by the live refresh lane"),
            ("HistoryMonths", 13, "number", "Closed-history retention window"),
        )
    elif name == "60_SOURCE_SYSTEMS":
        rows = SOURCE_ROLES
    elif name == "61_FIELD_MAPPING":
        rows = tuple(
            (
                "BLANK_DEPLOYMENT",
                "GenericDelimited",
                entity,
                "",
                canonical,
                adapter_data_type(canonical),
                canonical not in OPTIONAL_ADAPTER_FIELDS,
                False,
            )
            for entity, fields in ADAPTER_CONTRACTS.items()
            for canonical in fields
        )
    elif name == "71_METRIC_RULES":
        rows = (
            ("BLANK_DEPLOYMENT", "SERVICE_LEVEL", "Threshold", "", "seconds", "", "", False),
            ("BLANK_DEPLOYMENT", "SERVICE_LEVEL", "ExcludeShortAbandons", "", "boolean", "", "", False),
            ("BLANK_DEPLOYMENT", "ABANDON_RATE", "ShortAbandonThreshold", "", "seconds", "", "", False),
            ("BLANK_DEPLOYMENT", "ADHERENCE", "Tolerance", "", "minutes", "", "", False),
        )
    elif name == "72_FORECAST_POLICIES":
        rows = (
            (
                "BLANK_DEPLOYMENT", "DAILY_SEASONAL_BASELINE", "", "",
                "SEASONAL_NAIVE", "DAILY", 56, 28, 7, 28, "", "", False,
            ),
        )
    elif name == "73_CAPACITY_POLICIES":
        rows = (
            (
                "BLANK_DEPLOYMENT", "SYNCHRONOUS_STANDARD", "", "VOICE",
                "ERLANG_C", 30, 0.8, 20, 0.85, 0.2, 1, 1, "", "", False,
            ),
            (
                "BLANK_DEPLOYMENT", "ASYNCHRONOUS_STANDARD", "", "EMAIL",
                "WORKLOAD", 30, "", "", 0.85, 0.2, 1, 1, "", "", False,
            ),
        )
    style_table(ws, 13, 2, headers, rows, table_name, input_table=True, max_input_row=1000)
    end_col = get_column_letter(1 + len(headers))
    if "Enabled" in headers:
        col = get_column_letter(2 + headers.index("Enabled"))
        add_list_validation(ws, f"{col}14", f"{col}1000", "BOOLEAN_LIST", "Enable only after validation.")
    if "Approved" in headers:
        col = get_column_letter(2 + headers.index("Approved"))
        add_list_validation(ws, f"{col}14", f"{col}1000", "BOOLEAN_LIST", "Approve only after reconciliation.")
    if "ApprovalStatus" in headers:
        col = get_column_letter(2 + headers.index("ApprovalStatus"))
        add_list_validation(ws, f"{col}14", f"{col}1000", "APPROVAL_LIST", "Select the governed approval state.")
    if "Decision" in headers:
        col = get_column_letter(2 + headers.index("Decision"))
        add_list_validation(ws, f"{col}14", f"{col}1000", "ATTENDANCE_DECISION_LIST", "Select the attendance decision.")
    if "Severity" in headers:
        col = get_column_letter(2 + headers.index("Severity"))
        add_list_validation(ws, f"{col}14", f"{col}1000", "SEVERITY_LIST", "Select operational severity.")
    if "Status" in headers:
        col = get_column_letter(2 + headers.index("Status"))
        add_list_validation(ws, f"{col}14", f"{col}1000", "ACTION_STATUS_LIST", "Select the current lifecycle state.")
    channel_header = "ChannelKey" if "ChannelKey" in headers else "Channel" if "Channel" in headers else None
    if channel_header:
        col = get_column_letter(2 + headers.index(channel_header))
        add_list_validation(ws, f"{col}14", f"{col}1000", "CHANNEL_LIST", "Select a canonical channel.")
    if "Format" in headers:
        col = get_column_letter(2 + headers.index("Format"))
        add_list_validation(ws, f"{col}14", f"{col}1000", "FORMAT_LIST", "Choose the source extract format.")
    if "Method" in headers:
        col = get_column_letter(2 + headers.index("Method"))
        list_name = "CAPACITY_METHOD_LIST" if name == "73_CAPACITY_POLICIES" else "FORECAST_METHOD_LIST"
        add_list_validation(ws, f"{col}14", f"{col}1000", list_name, "Choose a governed planning method.")
    if "Frequency" in headers:
        col = get_column_letter(2 + headers.index("Frequency"))
        add_list_validation(ws, f"{col}14", f"{col}1000", "FORECAST_FREQUENCY_LIST", "Choose the implemented planning grain.")
    if "ImpactType" in headers:
        col = get_column_letter(2 + headers.index("ImpactType"))
        add_list_validation(ws, f"{col}14", f"{col}1000", "PLANNING_IMPACT_LIST", "Choose the governed impact domain.")
    for index, header in enumerate(headers, start=2):
        width = 14
        if header in {"Notes", "Reason", "DecisionReason", "EvidenceReference", "Outcome", "PathPattern", "TransformHint"}:
            width = 24
        elif "At" in header or "Date" in header or header in {"ValidFrom", "ValidTo", "Start", "End"}:
            width = 18
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.auto_filter.ref = None


def add_lookup_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("90_LOOKUPS")
    set_canvas(ws, "ink-500")
    lists = {
        "BOOLEAN_LIST": ("TRUE", "FALSE"),
        "PROFILE_LIST": ("BLANK_DEPLOYMENT",),
        "TIMEZONE_LIST": ("UTC", "Europe/Berlin", "Europe/London", "America/New_York", "America/Chicago", "America/Los_Angeles", "Asia/Dubai", "Asia/Kolkata", "Asia/Singapore", "Australia/Sydney"),
        "REFRESH_MODE_LIST": ("VALIDATION", "OPERATIONAL"),
        "DQ_STATUS_LIST": ("NOT RUN", "PASSED", "WARNING", "BLOCKING"),
        "APPROVAL_LIST": ("DRAFT", "PENDING", "APPROVED", "REJECTED", "SUPERSEDED"),
        "ATTENDANCE_DECISION_LIST": ("PENDING", "VALID", "EXCUSED", "UNEXCUSED", "CANCELLED"),
        "SEVERITY_LIST": ("INFORMATION", "WARNING", "BLOCKING"),
        "ACTION_STATUS_LIST": ("DRAFT", "OPEN", "IN PROGRESS", "BLOCKED", "COMPLETED", "CANCELLED"),
        "CHANNEL_LIST": ("VOICE", "CHAT", "EMAIL", "CASE", "BACK_OFFICE", "DISPATCH", "OTHER"),
        "FORMAT_LIST": ("Auto", "csv", "xlsx", "xlsb", "json", "parquet", "txt"),
        "FORECAST_METHOD_LIST": ("SEASONAL_NAIVE",),
        "CAPACITY_METHOD_LIST": ("ERLANG_C", "WORKLOAD"),
        "FORECAST_FREQUENCY_LIST": ("DAILY",),
        "PLANNING_IMPACT_LIST": ("VOLUME_PCT", "AHT_PCT", "SHRINKAGE_PCT"),
    }
    for col, (name, values) in enumerate(lists.items(), start=1):
        ws.cell(1, col, name).font = font(9, "ink-700", True)
        for row, value in enumerate(values, start=2):
            ws.cell(row, col, value).font = font(9, "ink-700")
        col_letter = get_column_letter(col)
        wb.defined_names.add(DefinedName(name, attr_text=f"'90_LOOKUPS'!${col_letter}$2:${col_letter}${len(values)+1}"))
        ws.column_dimensions[col_letter].width = max(16, max(map(len, values)) + 2)
    ws.sheet_state = "hidden"


def technical_sheet(wb: Workbook, name: str, title: str, purpose: str, status: str) -> None:
    ws = wb.create_sheet(name)
    set_canvas(ws, "ink-500")
    apply_fill_to_range(ws, 2, 3, 2, 17, "shell-950")
    merge_style(ws, "B2:E3", "WFM OS", cell_fill="shell-950", cell_font=font(18, "paper-000", True))
    merge_style(ws, "F2:Q3", title.upper(), cell_fill="shell-950", cell_font=font(9, "paper-000", True))
    merge_style(ws, "B5:Q7", purpose, cell_fill="paper-000", cell_font=font(10, "ink-700"), alignment=Alignment(vertical="center", wrap_text=True, indent=1), border=Border(left=side("ink-500", "medium")))
    merge_style(ws, "B9:Q12", status, cell_fill="warning-100", cell_font=font(10, "warning-600", True), alignment=Alignment(vertical="center", wrap_text=True, indent=1), border=Border(left=side("warning-600", "medium")))
    ws.sheet_state = "hidden"


def add_test_harness(wb: Workbook) -> None:
    ws = wb.create_sheet("98_TEST_HARNESS")
    set_canvas(ws, "ink-500")
    apply_fill_to_range(ws, 2, 3, 2, 17, "shell-950")
    merge_style(ws, "B2:E3", "WFM OS", cell_fill="shell-950", cell_font=font(18, "paper-000", True))
    merge_style(ws, "F2:Q3", "TECHNICAL TEST HARNESS", cell_fill="shell-950", cell_font=font(9, "paper-000", True))
    tests = (
        ("T-001", "Workbook opens", "Structure", "Workbook package opens without repair", "PASS", "Generator validation"),
        ("T-002", "Production data absent", "Security", "No production or inferred employee data", "PASS", "Data-free shell"),
        ("T-003", "Power Query executable", "Engine", "Required queries execute", "NOT IMPLEMENTED", "Desktop Excel release step"),
        ("T-004", "Power Pivot model executable", "Engine", "Tables, relationships, and measures execute", "NOT IMPLEMENTED", "Desktop Excel release step"),
        ("T-005", "Python in Excel executable", "Engine", "Analytical cells execute in declared order", "NOT IMPLEMENTED", "Desktop Excel release step"),
        ("T-006", "VBA controller executable", "Engine", "Signed refresh and publication controls execute", "NOT IMPLEMENTED", "Macro-enabled release step"),
        ("T-007", "Planning Python deterministic", "Planning", "Forecast, adjustment, accuracy, and capacity fixtures match", "PASS", "91_Tests/test_planning_cycle.py"),
    )
    style_table(ws, 6, 2, ("TestKey", "Test", "Domain", "Expected", "Status", "Evidence"), tests, "tblTestHarness")
    ws.sheet_state = "hidden"


def add_snapshot_store(wb: Workbook) -> None:
    ws = wb.create_sheet("94_SNAPSHOT_STORE")
    set_canvas(ws, "ink-500")
    apply_fill_to_range(ws, 2, 3, 2, 17, "shell-950")
    merge_style(ws, "B2:E3", "WFM OS", cell_fill="shell-950", cell_font=font(18, "paper-000", True))
    merge_style(ws, "F2:Q3", "CONTROLLED OPERATIONAL SNAPSHOTS", cell_fill="shell-950", cell_font=font(9, "paper-000", True))
    merge_style(
        ws,
        "B5:Q8",
        "Append-only store reserved for approved close-day snapshots. Do not type, paste, sort, or delete rows manually.",
        cell_fill="warning-100",
        cell_font=font(9, "warning-600", True),
        alignment=Alignment(vertical="center", wrap_text=True, indent=1),
        border=Border(left=side("warning-600", "medium")),
    )
    style_table(
        ws,
        10,
        2,
        (
            "SnapshotKey", "Profile", "BusinessDate", "IntervalStart", "ActivityKey",
            "ScheduledFTE", "ScheduledProductiveFTE", "PresentFTE", "ProductiveFTE",
            "RequiredFTE", "NetProductiveFTE", "Status", "ClosedAt", "ClosedBy", "SourceRunKey"
        ),
        [],
        "tblOperationalSnapshots",
    )
    ws.sheet_state = "hidden"


def add_query_outputs(wb: Workbook) -> None:
    ws = wb.create_sheet("95_QUERY_OUTPUTS")
    set_canvas(ws, "ink-500")
    apply_fill_to_range(ws, 2, 3, 2, 17, "shell-950")
    merge_style(ws, "B2:E3", "WFM OS", cell_fill="shell-950", cell_font=font(18, "paper-000", True))
    merge_style(ws, "F2:Q3", "BOUNDED QUERY OUTPUTS", cell_fill="shell-950", cell_font=font(9, "paper-000", True))
    merge_style(
        ws,
        "B5:Q8",
        "Reserved load surface for reconciled Power Query outputs. The current shell contains headers only; Power Query is not embedded.",
        cell_fill="info-100",
        cell_font=font(9, "info-600", True),
        alignment=Alignment(vertical="center", wrap_text=True, indent=1),
        border=Border(left=side("info-600", "medium")),
    )
    style_table(
        ws,
        10,
        2,
        (
            "SnapshotKey", "Profile", "BusinessDate", "IntervalStart", "ActivityKey",
            "ScheduledFTE", "ScheduledProductiveFTE", "PresentFTE", "ProductiveFTE",
            "RequiredFTE", "NetProductiveFTE", "Status"
        ),
        [],
        "tblCloseDayReady",
    )
    ws.sheet_state = "hidden"


def add_build_info(wb: Workbook, build_date: str, git_commit: str, version: str) -> None:
    ws = wb.create_sheet("99_BUILD_INFO")
    page_shell(ws, ws.title, "Build information", "Confirm exactly what this artifact contains before using it for an operational decision.", "info-600", "86_REQUIREMENT_APPROVAL", "00_HOME")
    rows = (
        ("Product", "WFM OS", "Universal, vendor-neutral Excel WFM application"),
        ("Artifact", "WFM_OS.xlsx", "Data-free application shell"),
        ("Release", version, "Semantic release identifier"),
        ("Build type", "APPLICATION SHELL", "Structure, design, tables, validation, and formulas"),
        ("Operational status", "NOT OPERATIONAL", "Do not use for workforce decisions yet"),
        ("Build date", build_date, "UTC date supplied to the generator"),
        ("Git commit", git_commit, "Source revision used for this build"),
        ("Canonical contracts", "1.2.0", "00_Governance/00-02_CANONICAL_CONTRACTS.md"),
        ("Design system", "1.0.0", "Obsidian & Pearl"),
        ("Power Query", "NOT EMBEDDED", "Must be installed and validated in desktop Excel"),
        ("Power Pivot / DAX", "NOT EMBEDDED", "Must be installed and validated in desktop Excel"),
        ("Python in Excel", "NOT EMBEDDED", "Must be installed and validated in desktop Excel"),
        ("VBA", "NOT EMBEDDED", "This .xlsx contains no macros"),
        ("Production data", "NONE", "Configuration defaults and empty tables only"),
        ("Generator", "tools/build_workbook.py", "Version-controlled source of this shell"),
    )
    merge_style(ws, "B7:Q10", "RELEASE GATE · The application is a premium executable shell only. Engine integration is deliberately not represented as complete.", cell_fill="warning-100", cell_font=font(9, "warning-600", True), alignment=Alignment(vertical="center", wrap_text=True, indent=1), border=Border(left=side("warning-600", "medium")))
    style_table(ws, 13, 2, ("Field", "Value", "Evidence / note"), rows, "tblBuildInfo")
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 58


def set_workbook_properties(wb: Workbook, build_date: str, version: str) -> None:
    stamp = datetime.fromisoformat(build_date).replace(tzinfo=timezone.utc)
    wb.properties.creator = "WFM OS"
    wb.properties.lastModifiedBy = "WFM OS deterministic generator"
    wb.properties.title = "WFM OS — Universal Workforce Management"
    wb.properties.subject = "Vendor-neutral workforce-management application shell"
    wb.properties.description = "Data-free Obsidian & Pearl WFM OS application shell"
    wb.properties.keywords = "WFM, Excel, workforce management, vendor neutral"
    wb.properties.category = "Business application"
    wb.properties.version = version
    wb.properties.created = stamp
    wb.properties.modified = stamp
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"


def normalize_zip_metadata(source: Path, destination: Path, build_date: str) -> None:
    """Write stable ZIP ordering and timestamps for reproducible OOXML output."""
    with zipfile.ZipFile(source, "r") as incoming, zipfile.ZipFile(destination, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9) as outgoing:
        for name in sorted(incoming.namelist()):
            data = incoming.read(name)
            if name == "docProps/core.xml":
                fixed_stamp = f"{build_date}T00:00:00Z".encode("ascii")
                data = re.sub(
                    rb"(<dcterms:modified[^>]*>).*?(</dcterms:modified>)",
                    rb"\g<1>" + fixed_stamp + rb"\g<2>",
                    data,
                )
            info = zipfile.ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = 0o600 << 16
            info.create_system = 3
            outgoing.writestr(info, data)


def current_commit(repo_root: Path) -> str:
    result = subprocess.run(
        ["git", "rev-parse", "--short=12", "HEAD"],
        cwd=repo_root,
        check=False,
        capture_output=True,
        text=True,
    )
    return result.stdout.strip() if result.returncode == 0 else "UNAVAILABLE"


def build(output: Path, build_date: str, git_commit: str, version: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    set_workbook_properties(wb, build_date, version)
    add_lookup_sheet(wb)
    add_home(wb, version)
    add_control_center(wb)
    add_data_quality(wb)
    add_refresh_audit(wb)

    cycle_names = ["03_REFRESH_AUDIT"] + [page.name for page in BUSINESS_PAGES] + ["60_SOURCE_SYSTEMS"]
    for index, page in enumerate(BUSINESS_PAGES, start=1):
        business_page(wb.create_sheet(page.name), page, cycle_names[index - 1], cycle_names[index + 1])

    for spec in CONFIG_SHEETS:
        config_or_input_sheet(wb, spec, is_input=False)
    for spec in INPUT_SHEETS:
        config_or_input_sheet(wb, spec, is_input=True)

    technical_sheet(wb, "91_PY_INIT", "Python initialization", "Reserved for shared imports, deterministic settings, and analytical execution order.", "NOT EMBEDDED IN THIS SHELL. Install genuine Python in Excel cells only in the desktop Excel release workflow.")
    technical_sheet(wb, "92_PY_FORECAST", "Python forecast lab", "Reserved for forecast backtesting and statistical candidate models fed by governed Excel or query outputs.", "NOT EMBEDDED IN THIS SHELL. No Python formula is claimed or simulated.")
    technical_sheet(wb, "93_PY_SCENARIOS", "Python scenario lab", "Reserved for simulation and optimization experiments; approved results must be versioned before use.", "NOT EMBEDDED IN THIS SHELL. No Python formula is claimed or simulated.")
    add_snapshot_store(wb)
    add_query_outputs(wb)
    technical_sheet(wb, "96_PIVOT_SUPPORT", "Pivot support", "Reserved for technical PivotTables that drive controlled business layouts and slicers.", "POWER PIVOT AND PIVOTTABLES ARE NOT EMBEDDED IN THIS SHELL.")
    technical_sheet(wb, "97_MODEL_REGISTRY", "Model registry", "Reserved for the installed query, table, relationship, measure, and refresh-lane inventory.", "THE SEMANTIC MODEL IS NOT EMBEDDED IN THIS SHELL.")
    add_test_harness(wb)
    add_build_info(wb, build_date, git_commit, version)

    # Correct navigation to the next numbered configuration/input page.
    visible_sequence = [ws.title for ws in wb.worksheets if ws.sheet_state == "visible"]
    for index, sheet_name in enumerate(visible_sequence):
        if sheet_name in {"00_HOME", "99_BUILD_INFO"}:
            continue
        ws = wb[sheet_name]
        previous_name = visible_sequence[index - 1]
        next_name = visible_sequence[index + 1] if index + 1 < len(visible_sequence) else "00_HOME"
        if ws["O2"].value == "PREVIOUS":
            ws["O2"].hyperlink = f"#'{previous_name}'!B2"
        if ws["P2"].value == "NEXT":
            ws["P2"].hyperlink = f"#'{next_name}'!B2"

    # Technical sheets precede BUILD_INFO physically but stay hidden.
    wb._sheets.sort(key=lambda ws: (
        0 if ws.title == "00_HOME" else
        1 if ws.title[:2].isdigit() else 2,
        int(ws.title[:2]) if ws.title[:2].isdigit() else 999,
        ws.title,
    ))
    wb.active = wb.sheetnames.index("00_HOME")

    output.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(prefix="wfm_os_build_") as temp_dir:
        raw = Path(temp_dir) / "raw.xlsx"
        normalized = Path(temp_dir) / "normalized.xlsx"
        wb.save(raw)
        normalize_zip_metadata(raw, normalized, build_date)
        output.write_bytes(normalized.read_bytes())

    # Fail the build if the resulting OOXML cannot be parsed by openpyxl.
    check = load_workbook(output, read_only=False, data_only=False)
    required = {"00_HOME", "01_CONTROL_CENTER", "02_DATA_QUALITY", "10_STRATEGIC_PLAN", "30_INTRADAY", "40_PERFORMANCE", "99_BUILD_INFO"}
    missing = sorted(required.difference(check.sheetnames))
    if missing:
        raise RuntimeError(f"Workbook validation failed; missing sheets: {missing}")
    if check["99_BUILD_INFO"]["C18"].value != "NOT OPERATIONAL":
        raise RuntimeError("Workbook validation failed; release gate is missing")
    check.close()


def parse_args() -> argparse.Namespace:
    default_root = Path(__file__).resolve().parents[1]
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=Path, default=default_root / "01_Application" / "WFM_OS.xlsx")
    parser.add_argument("--build-date", default=os.environ.get("SOURCE_DATE", datetime.now(timezone.utc).date().isoformat()))
    parser.add_argument("--git-commit", default=None)
    parser.add_argument("--version", default="0.3.0-shell")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    repo_root = Path(__file__).resolve().parents[1]
    git_commit = args.git_commit or current_commit(repo_root)
    datetime.fromisoformat(args.build_date)  # validate the CLI value early
    build(args.output.resolve(), args.build_date, git_commit, args.version)
    print(f"Built {args.output.resolve()}")


if __name__ == "__main__":
    main()
